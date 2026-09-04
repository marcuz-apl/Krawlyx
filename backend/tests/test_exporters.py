"""Tests for the streaming export pipeline (PRD §6.4 FR-EXP-04..08).

Each test instantiates an `Exporter` against a fresh `Job` row and a
real `ExportTarget` (no network). We use synthetic ORM rows instead of
the worker integration to keep these tests fast and deterministic.
"""

from __future__ import annotations

import csv
from datetime import UTC, datetime

from openpyxl import load_workbook
from sqlalchemy import select

from app.core.db import SessionLocal
from app.core.security import hash_password
from app.exporters import Exporter
from app.models import EngineInstance, ExportTarget, Job, JobResult, User
from app.models import Target as TargetRow
from app.services.engines import encrypt_config
from tests._helpers import make_user

# ---- helpers ----


def _make_job(*, user_id: int, engine_id: int) -> int:
    with SessionLocal() as db:
        job = Job(
            created_by_id=user_id,
            engine_id=engine_id,
            options={},
            notes="csv rollover test",
            status="queued",
        )
        db.add(job)
        db.commit()
        db.refresh(job)
        # A target so the slug has something interesting.
        db.add(
            TargetRow(job_id=job.id, url="https://example.com/seed", status="pending", attempts=0)
        )
        db.commit()
        return job.id


def _make_target(
    name: str,
    *,
    path: str,
    fmt: str = "csv",
    split_size_mb: int = 1,
    enabled: bool = True,
    runner_selectable: bool = True,
) -> int:
    with SessionLocal() as db:
        row = ExportTarget(
            name=name,
            mode="folder",
            path=path,
            file_format=fmt,  # mapped to the "format" column
            split_size_mb=split_size_mb,
            runner_selectable=runner_selectable,
            enabled=enabled,
        )
        db.add(row)
        db.commit()
        db.refresh(row)
        return row.id


def _make_engine() -> int:
    with SessionLocal() as db:
        eid = EngineInstance(
            name="exp-eng",
            type="playtrafi",
            config_encrypted=encrypt_config({}),
            pooled=True,
        )
        db.add(eid)
        db.commit()
        db.refresh(eid)
        return eid.id


def _user_id() -> int:
    with SessionLocal() as db:
        u = User(username="exp-user", password_hash=hash_password("x"), role="admin")
        db.add(u)
        db.commit()
        db.refresh(u)
        return u.id


def _result_row(target_id: int, *, title: str = "t", body: str = "x") -> int:
    with SessionLocal() as db:
        r = JobResult(
            target_id=target_id,
            final_url="https://example.com/",
            http_status=200,
            title=title,
            content_markdown=f"# {body}",
            content_text=body,
            links_json=[],
            metadata_json={},
            error=None,
            duration_ms=42,
            fetched_at=datetime.now(UTC),
        )
        db.add(r)
        db.commit()
        db.refresh(r)
        return r.id


# ---- CSV ----


def test_csv_writer_emits_utf8_sig_bom(tmp_path) -> None:
    make_user("admin", "admin")
    eid = _make_engine()
    jid = _make_job(user_id=_user_id(), engine_id=eid)
    tid = _make_target("csv-bom", path=str(tmp_path), fmt="csv")
    with SessionLocal() as db:
        job = db.get(Job, jid)
        target = db.get(ExportTarget, tid)
        ex = Exporter(job, target)
        ex.open()
        # Add a target + a result for the export row.
        with SessionLocal() as db2:
            t = db2.scalar(
                select(TargetRow).where(TargetRow.job_id == jid, TargetRow.status == "pending")
            )
        assert t is not None
        rid = _result_row(t.id, title="hello", body="world")
        with SessionLocal() as db3:
            r = db3.get(JobResult, rid)
            ex.write_result(r, source_url=t.url)
        ex.close()

    parts = sorted(tmp_path.glob("*_part*.csv"))
    assert parts, "no part files written"
    # First 3 bytes of part 1 must be the UTF-8 BOM.
    head = parts[0].read_bytes()[:3]
    assert head == b"\xef\xbb\xbf", f"expected BOM, got {head!r}"


def test_csv_rolls_over_at_byte_boundary(tmp_path) -> None:
    """With split_size_mb=1 and ~3 MB of rows, multiple parts are written.

    The orchestrator caps `content_text` at 500 chars and `title` at
    500 chars, so the largest per-row export is ~1 KB. 2000 rows
    ≈ 2 MB > 1 MB cap → 2+ parts.

    To keep the test fast we skip the per-row DB roundtrip by
    writing directly through the orchestrator with constructed
    `JobResult` and `Target` objects (the orchestrator doesn't read
    from the DB on `write_result`, it just copies fields).
    """
    from datetime import datetime

    from app.exporters.exporter import Exporter as _Ex
    from app.models import JobResult as JR
    from app.models import Target as TRow

    make_user("admin", "admin")
    eid = _make_engine()
    jid = _make_job(user_id=_user_id(), engine_id=eid)
    tid = _make_target("csv-roll", path=str(tmp_path), fmt="csv", split_size_mb=1)
    with SessionLocal() as db:
        job = db.get(Job, jid)
        target = db.get(ExportTarget, tid)
        ex = _Ex(job, target)
        ex.open()

    # Build 2000 synthetic rows.
    fixed_title = "x" * 500
    fixed_url = "https://example.com/r"
    for i in range(2000):
        # Construct an in-memory JobResult + Target — the orchestrator
        # only reads attributes, not relationships.
        t_row = TRow(
            id=1000 + i, job_id=jid, url=f"https://example.com/{i}", status="done", attempts=1
        )
        r_row = JR(
            id=10_000 + i,
            target_id=t_row.id,
            final_url=fixed_url,
            http_status=200,
            title=fixed_title,
            content_markdown="# h",
            content_text="body",
            links_json=[],
            metadata_json={},
            error=None,
            duration_ms=42,
            fetched_at=datetime.now(UTC),
        )
        ex.write_result(r_row, source_url=t_row.url)

    ex.close()

    parts = sorted(tmp_path.glob("*_part*.csv"))
    assert len(parts) >= 2, f"expected multiple parts, got {len(parts)}"
    # Every part under the limit (with a small tolerance for the last row).
    for p in parts:
        size = p.stat().st_size
        assert size <= 1.1 * 1024 * 1024, f"{p.name} too big: {size}"
    # Manifest exists and lists every part.
    manifests = list(tmp_path.glob("*_manifest.json"))
    assert len(manifests) == 1
    import json as _json

    manifest = _json.loads(manifests[0].read_text(encoding="utf-8"))
    assert manifest["job_id"] == jid
    assert len(manifest["parts"]) == len(parts)
    # Each part's row count from the manifest matches the file.
    for entry, path in zip(manifest["parts"], parts):
        with path.open(encoding="utf-8-sig", newline="") as fh:
            reader = csv.reader(fh)
            data_rows = sum(1 for _ in reader) - 1  # minus header
        assert data_rows == entry["rows"], (
            f"{path.name}: manifest says {entry['rows']} but file has {data_rows}"
        )


# ---- XLSX ----


def test_xlsx_writer_streams_rows(tmp_path) -> None:
    make_user("admin", "admin")
    eid = _make_engine()
    jid = _make_job(user_id=_user_id(), engine_id=eid)
    tid = _make_target("xlsx", path=str(tmp_path), fmt="xlsx", split_size_mb=1)
    with SessionLocal() as db:
        job = db.get(Job, jid)
        target = db.get(ExportTarget, tid)
        ex = Exporter(job, target)
        ex.open()

    # 50 small rows; the XLSX should comfortably fit in one part with
    # 1 MB cap.
    for i in range(50):
        with SessionLocal() as db:
            t = TargetRow(
                job_id=jid,
                url=f"https://example.com/x{i}",
                status="done",
                attempts=1,
            )
            db.add(t)
            db.commit()
            db.refresh(t)
        _result_row(t.id, title=f"x{i}", body="y")
        with SessionLocal() as db:
            r = db.scalar(select(JobResult).where(JobResult.target_id == t.id))
            ex.write_result(r, source_url=t.url)

    with SessionLocal() as db:
        ex.close()

    parts = sorted(tmp_path.glob("*_part*.xlsx"))
    assert len(parts) >= 1
    # openpyxl can read every part back.
    for p in parts:
        wb = load_workbook(p, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        # First row is the header.
        assert rows[0][0] == "target_id"
        wb.close()


# ---- Degraded path ----


def test_unwritable_path_marks_exporter_degraded(tmp_path) -> None:
    """FR-EXP-08: a target the worker can't write to becomes a no-op
    exporter; the job's status flips to `export_degraded`."""
    make_user("admin", "admin")
    eid = _make_engine()
    jid = _make_job(user_id=_user_id(), engine_id=eid)
    # Point at a path that mkdir will refuse (file in the way).
    blocker = tmp_path / "blocker"
    blocker.write_text("not a directory", encoding="utf-8")
    bad_path = blocker / "subdir"
    tid = _make_target("bad", path=str(bad_path), fmt="csv")
    with SessionLocal() as db:
        job = db.get(Job, jid)
        target = db.get(ExportTarget, tid)
        ex = Exporter(job, target)
        ex.open()
    assert ex.is_degraded
    assert (
        "mkdir" in (ex.degrade_reason or "").lower()
        or "not a directory" in (ex.degrade_reason or "").lower()
        or "file" in (ex.degrade_reason or "").lower()
    )
    # write_result / close are no-ops on a degraded exporter.
    ex.write_result(None, source_url="https://example.com/")  # type: ignore[arg-type]
    ex.close()
    assert not list(tmp_path.glob("*_part*.csv"))


# ---- Test endpoint ----


def test_endpoint_probe_writes_and_removes(tmp_path) -> None:
    from fastapi.testclient import TestClient

    from app.main import create_app
    from tests._helpers import auth_as

    # We need the export target to exist; create it then probe.
    tid = _make_target("probe", path=str(tmp_path), fmt="csv")

    with TestClient(create_app()) as c:
        csrf = auth_as(c, "u1", "admin")
        r = c.post(f"/api/export-targets/{tid}/test", headers={"X-CSRF-Token": csrf})
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["ok"] is True
        # No probe files left behind.
        assert not list(tmp_path.glob("Krawlyx_probe_*.txt"))


def test_endpoint_probe_reports_permission_failure(tmp_path) -> None:
    """A non-existent path makes the probe fail; the endpoint reports
    the error clearly (FR-EXP-08).

    We use a path inside a non-existent drive (`Z:\nope\\...`) on
    Windows, or `/this/path/should/not/exist/...` on POSIX. Either
    way the OS refuses to create the directory tree.
    """
    import sys

    from fastapi.testclient import TestClient

    from app.main import create_app
    from tests._helpers import auth_as

    if sys.platform == "win32":
        bad_path = "Z:\\Krawlyx-bad-target\\exports"
    else:
        bad_path = "/this/path/should/not/exist/Krawlyx-bad-target"

    tid = _make_target("bad", path=bad_path, fmt="csv")

    with TestClient(create_app()) as c:
        csrf = auth_as(c, "u2", "admin")
        r = c.post(f"/api/export-targets/{tid}/test", headers={"X-CSRF-Token": csrf})

    assert r.status_code == 200
    body = r.json()
    # The probe should report failure.
    assert body["ok"] is False
    assert body["detail"]  # surface the error message


def test_exporter_respects_custom_filename(tmp_path) -> None:
    """When job.options contains export_filename, the slug uses it."""
    user = make_user("exp_cust_fn")
    eid = _make_engine()
    with SessionLocal() as db:
        job = Job(
            created_by_id=user.id,
            engine_id=eid,
            options={"export_filename": "autotrader_sedans_2026"},
            notes="some notes",
            status="queued",
        )
        db.add(job)
        db.commit()
        db.refresh(job)
        jid = job.id

    tid = _make_target("custom_fn_target", path=str(tmp_path), fmt="csv")
    with SessionLocal() as db:
        job_row = db.get(Job, jid)
        target_row = db.get(ExportTarget, tid)
        exporter = Exporter(job_row, target_row)
        assert "autotrader_sedans_2026" in exporter._slug
